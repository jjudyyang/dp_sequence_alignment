"""Excel sequence-alignment engine used by the Shiftline web app."""

from copy import copy
from dataclasses import dataclass

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter


DEFAULT_INPUT_FILE = "input.xlsx"
DEFAULT_OUTPUT_FILE = "output.xlsx"
DEFAULT_INPUT_SHEET_NAME = "sheet1"
DEFAULT_OUTPUT_SHEET_NAME = "sheet 1 aligned result"
DEFAULT_START_ROW = 2
DEFAULT_HEADER_FIRST_ROW = 1
DEFAULT_HEADER_LAST_ROW = 1

DEFAULT_LEFT_INPUT_COL = "D"
DEFAULT_LEFT_BLOCK_START_COL = "A"
DEFAULT_LEFT_BLOCK_END_COL = "D"
DEFAULT_LEFT_OUTPUT_START_COL = "A"

DEFAULT_RIGHT_INPUT_COL = "F"
DEFAULT_RIGHT_BLOCK_START_COL = "F"
DEFAULT_RIGHT_BLOCK_END_COL = "H"
DEFAULT_RIGHT_OUTPUT_START_COL = "F"

DEFAULT_THRESHOLD = 0.5
DEFAULT_DIFF_OUTPUT_COL = "E"

MATCH_SCORE = 2
MISMATCH_SCORE = -10
GAP_SCORE = -1

UNMATCHED_FILL = PatternFill(fill_type="solid", fgColor="FFF9C4")
DIFF_ROW_FILL = PatternFill(fill_type="solid", fgColor="FFF4E6")


def to_float(value):
    """Convert an Excel cell value to a float, or None when it is blank/invalid."""
    try:
        if value is None or value == "":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def col_to_num(col_letter):
    """Convert an Excel column letter, such as A or AA, to a 1-based number."""
    return column_index_from_string(col_letter)


def normalize_column(value, field_name):
    """Validate and normalize an Excel column letter."""
    cleaned = (value or "").strip().upper()
    if not cleaned:
        raise ValueError(f"{field_name} is required.")
    try:
        col_to_num(cleaned)
    except ValueError:
        raise ValueError(
            f"{field_name} must be valid Excel column letters, for example A or AA."
        ) from None
    return cleaned


def normalize_header_text(value):
    """Return a compact lowercase header string for loose matching."""
    return " ".join(str(value or "").strip().lower().split())


def header_contains(value, *needles):
    """Return True when all requested words appear in a normalized cell value."""
    text = normalize_header_text(value)
    return bool(text) and all(needle in text for needle in needles)


def block_width_cols(start_col, end_col):
    """Return the width of an inclusive Excel column range."""
    return col_to_num(end_col) - col_to_num(start_col) + 1


def output_block_extent(output_start_col, block_start_col, block_end_col):
    """Return inclusive output column indexes where a pasted block will land."""
    start = col_to_num(output_start_col)
    end = start + block_width_cols(block_start_col, block_end_col) - 1
    return start, end


def output_block_end_col(output_start_col, block_start_col, block_end_col):
    """Return the last output column letter for a mapped block."""
    _, end = output_block_extent(output_start_col, block_start_col, block_end_col)
    return get_column_letter(end)


def sanitize_excel_sheet_title(title, default="Aligned"):
    """Return a worksheet name that is safe for Excel."""
    raw = (title or "").strip() or default
    illegal = "\\/*?:[]"
    cleaned = "".join(char for char in raw if char not in illegal).strip()
    return (cleaned or default)[:31]


def validate_match_column_in_block(match_col, block_start, block_end, side_label):
    """Ensure the match column is inside the copied column block."""
    match = col_to_num(match_col)
    start = col_to_num(block_start)
    end = col_to_num(block_end)
    if start > end:
        raise ValueError(
            f"{side_label}: block start {block_start} must be before or equal to {block_end}."
        )
    if not start <= match <= end:
        raise ValueError(
            f"{side_label}: match column {match_col} must be between "
            f"{block_start} and {block_end}."
        )


@dataclass(frozen=True)
class AlignmentConfig:
    """Normalized settings for one workbook alignment run."""

    input_sheet_name: str = DEFAULT_INPUT_SHEET_NAME
    output_sheet_name: str = DEFAULT_OUTPUT_SHEET_NAME
    start_row: int = DEFAULT_START_ROW
    header_first_row: int = DEFAULT_HEADER_FIRST_ROW
    header_last_row: int = DEFAULT_HEADER_LAST_ROW

    left_sheet_name: str | None = None
    right_sheet_name: str | None = None

    left_input_col: str = DEFAULT_LEFT_INPUT_COL
    left_block_start_col: str = DEFAULT_LEFT_BLOCK_START_COL
    left_block_end_col: str = DEFAULT_LEFT_BLOCK_END_COL
    left_output_start_col: str = DEFAULT_LEFT_OUTPUT_START_COL

    right_input_col: str = DEFAULT_RIGHT_INPUT_COL
    right_block_start_col: str = DEFAULT_RIGHT_BLOCK_START_COL
    right_block_end_col: str = DEFAULT_RIGHT_BLOCK_END_COL
    right_output_start_col: str = DEFAULT_RIGHT_OUTPUT_START_COL

    threshold: float = DEFAULT_THRESHOLD
    diff_output_col: str = DEFAULT_DIFF_OUTPUT_COL

    def __post_init__(self):
        object.__setattr__(
            self, "input_sheet_name", (self.input_sheet_name or "").strip()
        )
        object.__setattr__(
            self,
            "output_sheet_name",
            sanitize_excel_sheet_title(self.output_sheet_name),
        )
        object.__setattr__(self, "start_row", int(self.start_row))
        object.__setattr__(self, "header_first_row", int(self.header_first_row))
        object.__setattr__(self, "header_last_row", int(self.header_last_row))
        object.__setattr__(self, "threshold", float(self.threshold))
        object.__setattr__(
            self, "left_sheet_name", clean_optional_text(self.left_sheet_name)
        )
        object.__setattr__(
            self, "right_sheet_name", clean_optional_text(self.right_sheet_name)
        )

        if self.header_last_row < self.header_first_row:
            raise ValueError(
                "Header last row must be greater than or equal to header first row."
            )
        if self.start_row <= self.header_last_row:
            raise ValueError(
                "Data row must be below all header rows (after header last row)."
            )

        for field_name in (
            "left_input_col",
            "left_block_start_col",
            "left_block_end_col",
            "left_output_start_col",
            "right_input_col",
            "right_block_start_col",
            "right_block_end_col",
            "right_output_start_col",
        ):
            object.__setattr__(
                self,
                field_name,
                normalize_column(getattr(self, field_name), field_name),
            )

        validate_match_column_in_block(
            self.left_input_col,
            self.left_block_start_col,
            self.left_block_end_col,
            "Left column group",
        )
        validate_match_column_in_block(
            self.right_input_col,
            self.right_block_start_col,
            self.right_block_end_col,
            "Right column group",
        )

        right_output_start = nudge_right_output_if_overlaps(self)
        object.__setattr__(self, "right_output_start_col", right_output_start)

        diff_col = (self.diff_output_col or "").strip().upper()
        if diff_col:
            normalize_column(diff_col, "diff_output_col")
            validate_diff_column_clear_of_blocks(self, diff_col)
        object.__setattr__(self, "diff_output_col", diff_col)


def clean_optional_text(value):
    """Normalize optional form text into None or a stripped string."""
    cleaned = (value or "").strip()
    return cleaned or None


def nudge_right_output_if_overlaps(config):
    """Move the right output block after the left block if the two would overlap."""
    _, left_end = output_block_extent(
        config.left_output_start_col,
        config.left_block_start_col,
        config.left_block_end_col,
    )
    right_start = col_to_num(config.right_output_start_col)
    if right_start <= left_end:
        return get_column_letter(left_end + 1)
    return config.right_output_start_col


def validate_diff_column_clear_of_blocks(config, diff_col):
    """Difference column cannot sit inside either pasted output block."""
    diff = col_to_num(diff_col)
    left_start, left_end = output_block_extent(
        config.left_output_start_col,
        config.left_block_start_col,
        config.left_block_end_col,
    )
    right_start, right_end = output_block_extent(
        config.right_output_start_col,
        config.right_block_start_col,
        config.right_block_end_col,
    )
    if left_start <= diff <= left_end or right_start <= diff <= right_end:
        raise ValueError(
            f"Difference column {diff_col} overlaps pasted data. Choose a gap "
            "column between the left and right output blocks."
        )


def find_sheet_name(workbook, requested_name):
    """Find a worksheet by exact name, then by case-insensitive name."""
    if not requested_name:
        return None
    if requested_name in workbook.sheetnames:
        return requested_name
    requested_lower = requested_name.lower()
    for sheet_name in workbook.sheetnames:
        if sheet_name.lower() == requested_lower:
            return sheet_name
    return None


def resolve_sheet(workbook, requested_name):
    """Return the requested worksheet, or the active sheet when no name is provided."""
    matched_name = find_sheet_name(workbook, requested_name)
    if matched_name:
        return workbook[matched_name]
    if requested_name:
        print(f"Warning: sheet '{requested_name}' was not found.")
        print(f"Using active sheet instead: '{workbook.active.title}'")
    return workbook.active


def copy_cell(src_cell, dst_cell):
    """Copy an Excel cell value and visual style."""
    dst_cell.value = src_cell.value
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)


def copy_block(src_ws, dst_ws, src_row, dst_row, src_start_col, src_end_col, dst_start_col):
    """Copy a horizontal cell block from one worksheet row to another."""
    src_start = col_to_num(src_start_col)
    src_end = col_to_num(src_end_col)
    dst_start = col_to_num(dst_start_col)
    for offset, src_col in enumerate(range(src_start, src_end + 1)):
        copy_cell(
            src_ws.cell(row=src_row, column=src_col),
            dst_ws.cell(row=dst_row, column=dst_start + offset),
        )


def copy_mapped_column_widths(src_ws, dst_ws, src_start_col, src_end_col, dst_start_col):
    """Copy visible column widths from a source block to its output position."""
    src_start = col_to_num(src_start_col)
    src_end = col_to_num(src_end_col)
    dst_start = col_to_num(dst_start_col)
    for offset, src_col in enumerate(range(src_start, src_end + 1)):
        src_letter = get_column_letter(src_col)
        dst_letter = get_column_letter(dst_start + offset)
        width = src_ws.column_dimensions[src_letter].width
        if width is not None:
            dst_ws.column_dimensions[dst_letter].width = width


def highlight_block(ws, row, start_col, end_col, fill):
    """Apply a fill across an inclusive output column range."""
    for col in range(col_to_num(start_col), col_to_num(end_col) + 1):
        ws.cell(row=row, column=col).fill = fill


def read_column(ws, col_letter, start_row):
    """Read one matching column into row/value records used by the DP engine."""
    rows = []
    for row in range(start_row, ws.max_row + 1):
        value = to_float(ws[f"{col_letter}{row}"].value)
        if value is not None:
            rows.append({"row": row, "value": value})
    return rows


def require_match_values(values, ws, col_letter, side_label, start_row):
    """Reject mappings that point at an empty or non-numeric match column."""
    if values:
        return
    raise ValueError(
        f"No numeric values found in {side_label} match column {col_letter} "
        f"on sheet '{ws.title}' at or below row {start_row}. Choose the sheet "
        "and column that contain the values to compare."
    )


def infer_single_sheet_config(ws, config):
    """Infer the common Shiftline layout from visible workbook headers."""
    max_scan_rows = min(ws.max_row, 5)
    max_scan_cols = ws.max_column
    left_match = None
    right_match = None
    diff_col = None
    header_last_row = None

    for row in range(1, max_scan_rows + 1):
        for col in range(1, max_scan_cols + 1):
            value = ws.cell(row=row, column=col).value
            text = normalize_header_text(value)
            if not text:
                continue
            if "diff" in text or " dif" in f" {text}":
                diff_col = diff_col or col
            if header_contains(value, "target", "joint", "length"):
                left_match = left_match or col
                header_last_row = header_last_row or row

    if left_match is None:
        return None

    search_start = (diff_col or left_match) + 1
    for row in range(header_last_row or 1, max_scan_rows + 1):
        for col in range(search_start, max_scan_cols + 1):
            value = ws.cell(row=row, column=col).value
            text = normalize_header_text(value)
            if not text:
                continue
            if "joint" in text and "length" in text and "target" not in text:
                right_match = col
                header_last_row = row
                break
        if right_match is not None:
            break

    if right_match is None:
        return None

    if diff_col is None:
        diff_col = left_match + 1

    header_row = header_last_row or 1
    header_first_row = header_row
    if header_row > 1:
        previous_row_values = [
            normalize_header_text(ws.cell(row=header_row - 1, column=col).value)
            for col in range(1, max_scan_cols + 1)
        ]
        if any(value in {"current", "previous"} for value in previous_row_values):
            header_first_row = header_row - 1

    last_header_col = max(
        (
            col
            for col in range(1, max_scan_cols + 1)
            if ws.cell(row=header_row, column=col).value not in (None, "")
        ),
        default=max_scan_cols,
    )

    left_block_end = max(1, diff_col - 1)
    right_block_start = diff_col + 1
    return AlignmentConfig(
        input_sheet_name=config.input_sheet_name,
        output_sheet_name=config.output_sheet_name,
        start_row=header_row + 1,
        header_first_row=header_first_row,
        header_last_row=header_row,
        left_sheet_name=config.left_sheet_name,
        right_sheet_name=config.right_sheet_name,
        left_input_col=get_column_letter(left_match),
        left_block_start_col="A",
        left_block_end_col=get_column_letter(left_block_end),
        left_output_start_col="A",
        right_input_col=get_column_letter(right_match),
        right_block_start_col=get_column_letter(right_block_start),
        right_block_end_col=get_column_letter(last_header_col),
        right_output_start_col=get_column_letter(right_block_start),
        threshold=config.threshold,
        diff_output_col=get_column_letter(diff_col),
    )


def should_try_inferred_single_sheet_config(left_ws, right_ws, config):
    """Limit automatic remapping to same-sheet uploads using the default shape."""
    return (
        left_ws is right_ws
        and config.left_input_col == DEFAULT_LEFT_INPUT_COL
        and config.right_input_col == DEFAULT_RIGHT_INPUT_COL
        and config.left_block_start_col == DEFAULT_LEFT_BLOCK_START_COL
        and config.left_block_end_col == DEFAULT_LEFT_BLOCK_END_COL
        and config.right_block_start_col == DEFAULT_RIGHT_BLOCK_START_COL
        and config.right_block_end_col == DEFAULT_RIGHT_BLOCK_END_COL
        and config.diff_output_col == DEFAULT_DIFF_OUTPUT_COL
    )


def is_match(left_value, right_value, threshold=DEFAULT_THRESHOLD):
    """Return True when two numeric values are close enough to align."""
    return abs(left_value - right_value) <= threshold


def pair_score(left_item, right_item, threshold):
    """Score one possible left/right pairing."""
    if is_match(left_item["value"], right_item["value"], threshold):
        return MATCH_SCORE
    return MISMATCH_SCORE


def align_with_dp(left, right, threshold=DEFAULT_THRESHOLD):
    """
    Align two ordered numeric sequences while preserving row order.

    Each input item is a dict with:
    - row: original Excel row number
    - value: numeric value used for matching
    """
    n = len(left)
    m = len(right)
    dp = [[0] * (m + 1) for _ in range(n + 1)]
    trace = [[None] * (m + 1) for _ in range(n + 1)]

    for i in range(1, n + 1):
        dp[i][0] = dp[i - 1][0] + GAP_SCORE
        trace[i][0] = "UP"
    for j in range(1, m + 1):
        dp[0][j] = dp[0][j - 1] + GAP_SCORE
        trace[0][j] = "LEFT"

    for i in range(1, n + 1):
        for j in range(1, m + 1):
            left_item = left[i - 1]
            right_item = right[j - 1]
            diag = dp[i - 1][j - 1] + pair_score(left_item, right_item, threshold)
            up = dp[i - 1][j] + GAP_SCORE
            left_gap = dp[i][j - 1] + GAP_SCORE
            best = max(diag, up, left_gap)
            dp[i][j] = best

            if best == diag and is_match(left_item["value"], right_item["value"], threshold):
                trace[i][j] = "DIAG"
            elif best == up:
                trace[i][j] = "UP"
            elif best == left_gap:
                trace[i][j] = "LEFT"
            else:
                trace[i][j] = "DIAG"

    alignment = []
    i = n
    j = m
    while i > 0 or j > 0:
        direction = trace[i][j]
        if direction == "DIAG":
            left_item = left[i - 1]
            right_item = right[j - 1]
            alignment.append(
                {
                    "left": left_item,
                    "right": right_item,
                    "matched": is_match(left_item["value"], right_item["value"], threshold),
                }
            )
            i -= 1
            j -= 1
        elif direction == "UP":
            alignment.append({"left": left[i - 1], "right": None, "matched": False})
            i -= 1
        elif direction == "LEFT":
            alignment.append({"left": None, "right": right[j - 1], "matched": False})
            j -= 1
        else:
            break

    alignment.reverse()
    return alignment


def create_output_sheet(workbook, left_ws, right_ws, config):
    """Create the aligned output worksheet without modifying source sheets."""
    if config.output_sheet_name in workbook.sheetnames:
        del workbook[config.output_sheet_name]

    out_ws = workbook.create_sheet(config.output_sheet_name)

    for header_row in range(config.header_first_row, config.header_last_row + 1):
        copy_block(
            left_ws,
            out_ws,
            header_row,
            header_row,
            config.left_block_start_col,
            config.left_block_end_col,
            config.left_output_start_col,
        )
        copy_block(
            right_ws,
            out_ws,
            header_row,
            header_row,
            config.right_block_start_col,
            config.right_block_end_col,
            config.right_output_start_col,
        )

    copy_mapped_column_widths(
        left_ws,
        out_ws,
        config.left_block_start_col,
        config.left_block_end_col,
        config.left_output_start_col,
    )
    copy_mapped_column_widths(
        right_ws,
        out_ws,
        config.right_block_start_col,
        config.right_block_end_col,
        config.right_output_start_col,
    )
    return out_ws


def highlight_mapped_block(ws, row, output_start_col, block_start_col, block_end_col):
    """Highlight the output columns occupied by a mapped source block."""
    highlight_block(
        ws,
        row,
        output_start_col,
        output_block_end_col(output_start_col, block_start_col, block_end_col),
        UNMATCHED_FILL,
    )


def write_alignment(left_ws, right_ws, out_ws, alignment, config):
    """Write aligned row blocks into the output worksheet."""
    output_row = config.start_row
    for step in alignment:
        left_item = step["left"]
        right_item = step["right"]
        matched = step["matched"]

        if left_item is not None:
            copy_block(
                left_ws,
                out_ws,
                left_item["row"],
                output_row,
                config.left_block_start_col,
                config.left_block_end_col,
                config.left_output_start_col,
            )

        if right_item is not None:
            copy_block(
                right_ws,
                out_ws,
                right_item["row"],
                output_row,
                config.right_block_start_col,
                config.right_block_end_col,
                config.right_output_start_col,
            )

        if not matched:
            if left_item is not None:
                highlight_mapped_block(
                    out_ws,
                    output_row,
                    config.left_output_start_col,
                    config.left_block_start_col,
                    config.left_block_end_col,
                )
            if right_item is not None:
                highlight_mapped_block(
                    out_ws,
                    output_row,
                    config.right_output_start_col,
                    config.right_block_start_col,
                    config.right_block_end_col,
                )

        output_row += 1


def write_difference_column(out_ws, alignment, config):
    """Fill the absolute-difference column for rows with both left and right values."""
    if not config.diff_output_col:
        return

    diff_col = col_to_num(config.diff_output_col)
    out_ws.cell(row=config.header_last_row, column=diff_col).value = "Abs diff"

    output_row = config.start_row
    for step in alignment:
        left_item = step["left"]
        right_item = step["right"]
        cell = out_ws.cell(row=output_row, column=diff_col)
        if left_item is not None and right_item is not None:
            cell.value = abs(left_item["value"] - right_item["value"])
            cell.fill = DIFF_ROW_FILL
        else:
            cell.value = None
        output_row += 1


def process_excel(
    input_file,
    output_file,
    input_sheet_name=DEFAULT_INPUT_SHEET_NAME,
    output_sheet_name=DEFAULT_OUTPUT_SHEET_NAME,
    start_row=DEFAULT_START_ROW,
    header_first_row=DEFAULT_HEADER_FIRST_ROW,
    header_last_row=DEFAULT_HEADER_LAST_ROW,
    left_input_col=DEFAULT_LEFT_INPUT_COL,
    left_block_start_col=DEFAULT_LEFT_BLOCK_START_COL,
    left_block_end_col=DEFAULT_LEFT_BLOCK_END_COL,
    left_output_start_col=DEFAULT_LEFT_OUTPUT_START_COL,
    right_input_col=DEFAULT_RIGHT_INPUT_COL,
    right_block_start_col=DEFAULT_RIGHT_BLOCK_START_COL,
    right_block_end_col=DEFAULT_RIGHT_BLOCK_END_COL,
    right_output_start_col=DEFAULT_RIGHT_OUTPUT_START_COL,
    left_sheet_name=None,
    right_sheet_name=None,
    threshold=DEFAULT_THRESHOLD,
    diff_output_col=DEFAULT_DIFF_OUTPUT_COL,
    verbose=False,
):
    """Align two spreadsheet sequences and save a new workbook."""
    config = AlignmentConfig(
        input_sheet_name=input_sheet_name,
        output_sheet_name=output_sheet_name,
        start_row=start_row,
        header_first_row=header_first_row,
        header_last_row=header_last_row,
        left_sheet_name=left_sheet_name,
        right_sheet_name=right_sheet_name,
        left_input_col=left_input_col,
        left_block_start_col=left_block_start_col,
        left_block_end_col=left_block_end_col,
        left_output_start_col=left_output_start_col,
        right_input_col=right_input_col,
        right_block_start_col=right_block_start_col,
        right_block_end_col=right_block_end_col,
        right_output_start_col=right_output_start_col,
        threshold=threshold,
        diff_output_col=diff_output_col,
    )

    workbook = load_workbook(input_file)
    left_ws = resolve_sheet(workbook, config.left_sheet_name or config.input_sheet_name)
    right_ws = resolve_sheet(
        workbook,
        config.right_sheet_name or config.left_sheet_name or config.input_sheet_name,
    )

    left_values = read_column(left_ws, config.left_input_col, config.start_row)
    right_values = read_column(right_ws, config.right_input_col, config.start_row)
    if (
        (not left_values or not right_values)
        and should_try_inferred_single_sheet_config(left_ws, right_ws, config)
    ):
        inferred_config = infer_single_sheet_config(left_ws, config)
        if inferred_config is not None:
            config = inferred_config
            left_values = read_column(left_ws, config.left_input_col, config.start_row)
            right_values = read_column(right_ws, config.right_input_col, config.start_row)
    require_match_values(
        left_values, left_ws, config.left_input_col, "left", config.start_row
    )
    require_match_values(
        right_values, right_ws, config.right_input_col, "right", config.start_row
    )
    alignment = align_with_dp(left_values, right_values, config.threshold)

    out_ws = create_output_sheet(workbook, left_ws, right_ws, config)
    write_alignment(left_ws, right_ws, out_ws, alignment, config)
    write_difference_column(out_ws, alignment, config)
    workbook.save(output_file)

    result = {
        "output_file": str(output_file),
        "left_sheet": left_ws.title,
        "right_sheet": right_ws.title,
        "output_sheet": config.output_sheet_name,
        "left_values": len(left_values),
        "right_values": len(right_values),
        "alignment_steps": len(alignment),
        "matches_written": sum(1 for step in alignment if step["matched"]),
    }
    if verbose:
        print("Done:", result["output_file"])
        print("Left sheet:", result["left_sheet"])
        print("Right sheet:", result["right_sheet"])
        print("Output sheet:", result["output_sheet"])
        print("Left values:", result["left_values"])
        print("Right values:", result["right_values"])
        print("Alignment steps:", result["alignment_steps"])
        print("Matches written:", result["matches_written"])
    return result


def main():
    """Run the default workbook alignment from the terminal."""
    process_excel(
        input_file=DEFAULT_INPUT_FILE,
        output_file=DEFAULT_OUTPUT_FILE,
        input_sheet_name=DEFAULT_INPUT_SHEET_NAME,
        output_sheet_name=DEFAULT_OUTPUT_SHEET_NAME,
        start_row=DEFAULT_START_ROW,
        header_first_row=DEFAULT_HEADER_FIRST_ROW,
        header_last_row=DEFAULT_HEADER_LAST_ROW,
        left_input_col=DEFAULT_LEFT_INPUT_COL,
        left_block_start_col=DEFAULT_LEFT_BLOCK_START_COL,
        left_block_end_col=DEFAULT_LEFT_BLOCK_END_COL,
        left_output_start_col=DEFAULT_LEFT_OUTPUT_START_COL,
        right_input_col=DEFAULT_RIGHT_INPUT_COL,
        right_block_start_col=DEFAULT_RIGHT_BLOCK_START_COL,
        right_block_end_col=DEFAULT_RIGHT_BLOCK_END_COL,
        right_output_start_col=DEFAULT_RIGHT_OUTPUT_START_COL,
        threshold=DEFAULT_THRESHOLD,
        diff_output_col=DEFAULT_DIFF_OUTPUT_COL,
        verbose=True,
    )


if __name__ == "__main__":
    main()
