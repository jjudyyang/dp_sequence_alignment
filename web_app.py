from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Any, Optional

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# =========================
# Defaults for your test file
# =========================

DEFAULTS = {
    "sheet_name": "orginal data",  # exact spelling from your message
    "first_series_compare_col": "D",
    "first_series_col_from": "A",
    "first_series_col_to": "D",
    "first_series_row_start": "2",
    "first_series_row_end": "",  # blank = use ws.max_row
    "second_series_compare_col": "F",
    "second_series_col_from": "F",
    "second_series_col_to": "H",
    "second_series_row_start": "2",
    "second_series_row_end": "",  # blank = use ws.max_row
    "threshold": "0.5",
}

MATCH_SCORE = 2
MISMATCH_SCORE = -10
GAP_SCORE = -1
UNMATCHED_FILL = PatternFill(fill_type="solid", fgColor="FFA500")

app = FastAPI(title="Shift Algorithm App")
templates = Jinja2Templates(directory=str(Path(__file__).parent / "templates"))


def to_float(value: Any):
    try:
        if value is None or value == "":
            return None
        return float(value)
    except Exception:
        return None


def parse_positive_int(value: Any, default: int, field_name: str) -> int:
    """Parse form values. Blank means use the provided default."""
    text = "" if value is None else str(value).strip()
    if text == "":
        return default
    try:
        parsed = int(text)
    except ValueError as exc:
        raise ValueError(f"{field_name} must be a whole number.") from exc
    if parsed < 1:
        raise ValueError(f"{field_name} must be at least 1.")
    return parsed


def parse_optional_positive_int(value: Any, field_name: str) -> Optional[int]:
    """Parse optional end-row values. Blank or 0 means use ws.max_row."""
    text = "" if value is None else str(value).strip()
    if text == "" or text == "0":
        return None
    try:
        parsed = int(text)
    except ValueError as exc:
        raise ValueError(f"{field_name} must be a whole number, blank, or 0.") from exc
    if parsed < 1:
        raise ValueError(f"{field_name} must be at least 1, blank, or 0.")
    return parsed


def normalize_col(col: str) -> str:
    col = (col or "").strip().upper()
    if not col.isalpha():
        raise ValueError("Column must be letters only, like A or AB.")
    return col


def col_span_sorted(col_from: str, col_to: str) -> tuple[int, int]:
    """Return 1-based inclusive column indices lo..hi after sorting."""
    a = column_index_from_string(normalize_col(col_from))
    b = column_index_from_string(normalize_col(col_to))
    lo, hi = (a, b) if a <= b else (b, a)
    return lo, hi


def assert_compare_in_span(compare_col: str, span_lo: int, span_hi: int, label: str) -> None:
    idx = column_index_from_string(normalize_col(compare_col))
    if not span_lo <= idx <= span_hi:
        raise ValueError(
            f"{label} compare column {get_column_letter(idx)} must fall inside "
            f"{get_column_letter(span_lo)}:{get_column_letter(span_hi)}."
        )


def read_column_in_range(ws: Worksheet, col_letter: str, start_row: int, end_row: int):
    rows = []
    for row in range(start_row, end_row + 1):
        value = to_float(ws[f"{col_letter}{row}"].value)
        if value is not None:
            rows.append({"row": row, "value": value})
    return rows


def is_match(left_value: float, right_value: float, threshold: float) -> bool:
    return abs(left_value - right_value) <= threshold


def pair_score(left_item, right_item, threshold: float) -> int:
    if is_match(left_item["value"], right_item["value"], threshold):
        return MATCH_SCORE
    return MISMATCH_SCORE


def align_with_dp(left, right, threshold: float):
    """
    Full sequence alignment.

    DIAG = align one first-series value with one second-series value.
    UP = first-series row has no matching second-series row.
    LEFT = second-series row has no matching first-series row.
    """
    n = len(left)
    m = len(right)

    dp = [[0] * (m + 1) for _ in range(n + 1)]
    trace = [[None] * (m + 1) for _ in range(n + 1)]

    for i in range(1, n + 1):
        dp[i][0] = dp[i - 1][0] + GAP_SCORE
        trace[i][0] = "UP"

    for j in range(1, m + 1):
        dp[0][j] = dp[0][j - 1] + GAP_SCORE
        trace[0][j] = "LEFT"

    for i in range(1, n + 1):
        for j in range(1, m + 1):
            left_item = left[i - 1]
            right_item = right[j - 1]

            diag = dp[i - 1][j - 1] + pair_score(left_item, right_item, threshold)
            up = dp[i - 1][j] + GAP_SCORE
            left_gap = dp[i][j - 1] + GAP_SCORE

            best = max(diag, up, left_gap)
            dp[i][j] = best

            # Prefer true matches. Mismatches are heavily penalized, so rows usually
            # become LEFT_ONLY / RIGHT_ONLY instead of being forced together.
            if best == diag and is_match(left_item["value"], right_item["value"], threshold):
                trace[i][j] = "DIAG"
            elif best == up:
                trace[i][j] = "UP"
            elif best == left_gap:
                trace[i][j] = "LEFT"
            else:
                trace[i][j] = "DIAG"

    alignment = []
    i = n
    j = m

    while i > 0 or j > 0:
        direction = trace[i][j]

        if direction == "DIAG":
            left_item = left[i - 1]
            right_item = right[j - 1]
            alignment.append(
                {
                    "left": left_item,
                    "right": right_item,
                    "matched": is_match(left_item["value"], right_item["value"], threshold),
                }
            )
            i -= 1
            j -= 1
        elif direction == "UP":
            left_item = left[i - 1]
            alignment.append({"left": left_item, "right": None, "matched": False})
            i -= 1
        elif direction == "LEFT":
            right_item = right[j - 1]
            alignment.append({"left": None, "right": right_item, "matched": False})
            j -= 1
        else:
            break

    alignment.reverse()
    return alignment


def merge_match_row_snapshots(
    original_rows,
    lr: int,
    rr: int,
    left_compare_idx: int,
    right_compare_idx: int,
    left_val: float,
    right_val: float,
    first_span_lo: int,
    first_span_hi: int,
    second_span_lo: int,
    second_span_hi: int,
    base_fallback_row: int,
) -> list[Any]:
    """
    Build one output row.

    For a matched pair:
    - columns in first span come from the first-series row
    - columns in second span come from the second-series row
    - all other columns come from the later/base row
    """
    max_len = len(original_rows[base_fallback_row])

    overlap_lo = max(first_span_lo, second_span_lo)
    overlap_hi = min(first_span_hi, second_span_hi)
    if overlap_lo <= overlap_hi:
        cols = ", ".join(get_column_letter(c) for c in range(overlap_lo, overlap_hi + 1))
        raise ValueError(f"First and second column spans overlap at {cols}; use disjoint ranges.")

    row = [None] * max_len

    left_row_src = original_rows.get(lr)
    right_row_src = original_rows.get(rr)
    base_row_src = original_rows[base_fallback_row]

    if left_row_src is None or len(left_row_src) < max_len:
        raise ValueError(f"Missing row snapshot for first-series row {lr}.")
    if right_row_src is None or len(right_row_src) < max_len:
        raise ValueError(f"Missing row snapshot for second-series row {rr}.")

    for c in range(1, max_len + 1):
        if first_span_lo <= c <= first_span_hi:
            row[c - 1] = left_row_src[c - 1]
        elif second_span_lo <= c <= second_span_hi:
            row[c - 1] = right_row_src[c - 1]
        else:
            row[c - 1] = base_row_src[c - 1]

    row[left_compare_idx - 1] = left_val
    row[right_compare_idx - 1] = right_val
    return row


def orange_fill_series_span(ws, row_idx: int, span_lo: int, span_hi: int) -> None:
    for col in range(span_lo, span_hi + 1):
        ws.cell(row=row_idx, column=col).fill = UNMATCHED_FILL


def process_workbook(
    source_bytes: bytes,
    sheet_name: str,
    first_series_compare_col: str,
    first_series_col_from: str,
    first_series_col_to: str,
    first_series_row_start: int,
    first_series_row_end: Optional[int],
    second_series_compare_col: str,
    second_series_col_from: str,
    second_series_col_to: str,
    second_series_row_start: int,
    second_series_row_end: Optional[int],
    threshold: float,
) -> tuple[bytes, dict[str, int]]:
    col_left = normalize_col(first_series_compare_col)
    col_right = normalize_col(second_series_compare_col)

    first_span_lo, first_span_hi = col_span_sorted(first_series_col_from, first_series_col_to)
    second_span_lo, second_span_hi = col_span_sorted(second_series_col_from, second_series_col_to)

    wb = load_workbook(BytesIO(source_bytes))
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}")
    ws = wb[sheet_name]

    if first_series_row_end is None:
        first_series_row_end = ws.max_row
    if second_series_row_end is None:
        second_series_row_end = ws.max_row

    if first_series_row_start < 1 or first_series_row_end < first_series_row_start:
        raise ValueError("First series row range is invalid (need start >= 1 and end >= start).")
    if second_series_row_start < 1 or second_series_row_end < second_series_row_start:
        raise ValueError("Second series row range is invalid (need start >= 1 and end >= start).")

    assert_compare_in_span(col_left, first_span_lo, first_span_hi, "First series")
    assert_compare_in_span(col_right, second_span_lo, second_span_hi, "Second series")

    left = read_column_in_range(ws, col_left, first_series_row_start, first_series_row_end)
    right = read_column_in_range(ws, col_right, second_series_row_start, second_series_row_end)
    alignment = align_with_dp(left, right, threshold)

    col_compare_left_idx = column_index_from_string(col_left)
    col_compare_right_idx = column_index_from_string(col_right)
    max_col = max(
        ws.max_column or 1,
        first_span_hi,
        second_span_hi,
        col_compare_left_idx,
        col_compare_right_idx,
    )

    rows_to_snapshot = sorted(
        set(range(first_series_row_start, first_series_row_end + 1))
        | set(range(second_series_row_start, second_series_row_end + 1))
    )
    original_rows = {
        row: [copy(ws.cell(row=row, column=col).value) for col in range(1, max_col + 1)]
        for row in rows_to_snapshot
    }

    write_start_row = min(first_series_row_start, second_series_row_start)
    baseline_end_before_stack = max(first_series_row_end, second_series_row_end)

    shifted_rows = []
    for step in alignment:
        left_item = step["left"]
        right_item = step["right"]
        matched = step["matched"]

        if left_item is not None and right_item is not None:
            if not matched:
                continue
            lr = left_item["row"]
            rr = right_item["row"]
            base_row = max(lr, rr)
            row_values = merge_match_row_snapshots(
                original_rows,
                lr,
                rr,
                col_compare_left_idx,
                col_compare_right_idx,
                left_item["value"],
                right_item["value"],
                first_span_lo,
                first_span_hi,
                second_span_lo,
                second_span_hi,
                base_row,
            )
            shifted_rows.append((row_values, "MATCH"))
        elif left_item is not None:
            shifted_rows.append((list(original_rows[left_item["row"]]), "LEFT_ONLY"))
        elif right_item is not None:
            shifted_rows.append((list(original_rows[right_item["row"]]), "RIGHT_ONLY"))

    # Output may extend past the original series ranges if stacking needs more rows.
    out_end_row = max(baseline_end_before_stack, write_start_row + len(shifted_rows) - 1)

    # Clear the output area, but leave rows above write_start_row alone.
    for row in range(write_start_row, out_end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)

    for idx, (row_values, status) in enumerate(shifted_rows):
        out_row = write_start_row + idx
        for col, value in enumerate(row_values, start=1):
            ws.cell(row=out_row, column=col).value = value
        if status == "LEFT_ONLY":
            orange_fill_series_span(ws, out_row, first_span_lo, first_span_hi)
        elif status == "RIGHT_ONLY":
            orange_fill_series_span(ws, out_row, second_span_lo, second_span_hi)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue(), {
        "left_values": len(left),
        "right_values": len(right),
        "alignment_steps": len(alignment),
        "matches_written": sum(
            1
            for step in alignment
            if step["left"] is not None
            and step["right"] is not None
            and step["matched"]
        ),
        "output_rows": len(shifted_rows),
    }


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "defaults": DEFAULTS,
        },
    )


@app.post("/process")
async def process_file(
    upload: UploadFile = File(...),
    sheet_name: str = Form(DEFAULTS["sheet_name"]),
    first_series_compare_col: str = Form(DEFAULTS["first_series_compare_col"]),
    first_series_col_from: str = Form(DEFAULTS["first_series_col_from"]),
    first_series_col_to: str = Form(DEFAULTS["first_series_col_to"]),
    first_series_row_start: Optional[str] = Form(DEFAULTS["first_series_row_start"]),
    first_series_row_end: Optional[str] = Form(DEFAULTS["first_series_row_end"]),
    second_series_compare_col: str = Form(DEFAULTS["second_series_compare_col"]),
    second_series_col_from: str = Form(DEFAULTS["second_series_col_from"]),
    second_series_col_to: str = Form(DEFAULTS["second_series_col_to"]),
    second_series_row_start: Optional[str] = Form(DEFAULTS["second_series_row_start"]),
    second_series_row_end: Optional[str] = Form(DEFAULTS["second_series_row_end"]),
    threshold: str = Form(DEFAULTS["threshold"]),
):
    if not upload.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported for now.")

    payload = await upload.read()

    try:
        first_start = parse_positive_int(first_series_row_start, 2, "First series start row")
        first_end = parse_optional_positive_int(first_series_row_end, "First series end row")
        second_start = parse_positive_int(second_series_row_start, 2, "Second series start row")
        second_end = parse_optional_positive_int(second_series_row_end, "Second series end row")
        threshold_value = float(str(threshold).strip())

        processed_bytes, stats = process_workbook(
            source_bytes=payload,
            sheet_name=sheet_name.strip() or DEFAULTS["sheet_name"],
            first_series_compare_col=first_series_compare_col,
            first_series_col_from=first_series_col_from,
            first_series_col_to=first_series_col_to,
            first_series_row_start=first_start,
            first_series_row_end=first_end,
            second_series_compare_col=second_series_compare_col,
            second_series_col_from=second_series_col_from,
            second_series_col_to=second_series_col_to,
            second_series_row_start=second_start,
            second_series_row_end=second_end,
            threshold=threshold_value,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    output_name = f"shifted_{Path(upload.filename).name}"
    headers = {
        "X-Left-Values": str(stats["left_values"]),
        "X-Right-Values": str(stats["right_values"]),
        "X-Alignment-Steps": str(stats["alignment_steps"]),
        "X-Matches-Written": str(stats["matches_written"]),
        "X-Output-Rows": str(stats["output_rows"]),
    }

    return StreamingResponse(
        BytesIO(processed_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            **headers,
            "Content-Disposition": f'attachment; filename="{output_name}"',
        },
    )
