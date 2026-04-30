from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter


# =========================
# Config
# =========================

INPUT_FILE = "input.xlsx"
OUTPUT_FILE = "output.xlsx"

INPUT_SHEET_NAME = "sheet1"

OUTPUT_SHEET_NAME = "sheet 1 aligned result"

START_ROW = 2

# Rows on the sheet that belong to headers (copied verbatim to output). Inclusive range.
HEADER_FIRST_ROW = 1
HEADER_LAST_ROW = 1

# Left side: match using D, move A:D together
LEFT_INPUT_COL = "D"
LEFT_BLOCK_START_COL = "A"
LEFT_BLOCK_END_COL = "D"
LEFT_OUTPUT_START_COL = "A"

# Right side: match using F, move F:H together
RIGHT_INPUT_COL = "F"
RIGHT_BLOCK_START_COL = "F"
RIGHT_BLOCK_END_COL = "H"
RIGHT_OUTPUT_START_COL = "F"

# Matching rule
THRESHOLD = 0.5

# After alignment: |left match − right match| is written here (must not overlap left/right output blocks).
DIFF_OUTPUT_COL = "E"


def output_block_extent(output_start_col, block_start_col, block_end_col):
    """Inclusive output column indexes where a pasted block occupies space."""
    w = block_width_cols(block_start_col, block_end_col)
    s = col_to_num(output_start_col)
    return s, s + w - 1


def assert_diff_column_clear_of_blocks(diff_col_letter):
    """Difference column cannot sit inside the pasted left or right regions."""
    d = col_to_num(diff_col_letter)
    ls, le = output_block_extent(
        LEFT_OUTPUT_START_COL, LEFT_BLOCK_START_COL, LEFT_BLOCK_END_COL
    )
    rs, re_end = output_block_extent(
        RIGHT_OUTPUT_START_COL, RIGHT_BLOCK_START_COL, RIGHT_BLOCK_END_COL
    )
    if ls <= d <= le or rs <= d <= re_end:
        raise ValueError(
            f"Difference column {diff_col_letter} overlaps pasted data. Choose a gap column "
            f"(between the left block and right block on the output sheet)."
        )



# DP scoring
MATCH_SCORE = 2
MISMATCH_SCORE = -10
GAP_SCORE = -1

# Highlight unmatched rows yellow
UNMATCHED_FILL = PatternFill(fill_type="solid", fgColor="FFF9C4")

# Highlight rows that have a computed abs-difference (very light orange)
DIFF_ROW_FILL = PatternFill(fill_type="solid", fgColor="FFF4E6")


# =========================
# Helpers
# =========================

def to_float(value):
    """
    Convert Excel value to float.
    Empty or invalid cells become None.
    """
    try:
        if value is None or value == "":
            return None
        return float(value)
    except Exception:
        return None


def col_to_num(col_letter):
    """
    Convert Excel column letter to number.

    Example:
        A -> 1
        D -> 4
        F -> 6
    """
    return column_index_from_string(col_letter)


def copy_cell(src_cell, dst_cell):
    """
    Copy cell value and style from one cell to another.
    """
    dst_cell.value = src_cell.value

    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)


def copy_block(src_ws, dst_ws, src_row, dst_row, src_start_col, src_end_col, dst_start_col):
    """
    Copy a horizontal block from one row to another.

    Example:
        Copy A3:D3 into A8:D8
    """
    src_start = col_to_num(src_start_col)
    src_end = col_to_num(src_end_col)
    dst_start = col_to_num(dst_start_col)

    for offset, src_col in enumerate(range(src_start, src_end + 1)):
        src_cell = src_ws.cell(row=src_row, column=src_col)
        dst_cell = dst_ws.cell(row=dst_row, column=dst_start + offset)
        copy_cell(src_cell, dst_cell)


def block_width_cols(start_col, end_col):
    """Number of columns in an inclusive A..Z / AA.. range."""
    return col_to_num(end_col) - col_to_num(start_col) + 1


def assert_match_column_in_block(match_col, block_start, block_end, side_label):
    """Match column must lie inside the block that moves together (e.g. A–F with match F)."""
    m = col_to_num(match_col)
    lo = col_to_num(block_start)
    hi = col_to_num(block_end)
    if lo > hi:
        raise ValueError(
            f"{side_label}: block start {block_start} must be at or before end {block_end}."
        )
    if not (lo <= m <= hi):
        raise ValueError(
            f"{side_label}: match column {match_col} must be between {block_start} and {block_end} "
            f"(it should be one of the columns you are copying)."
        )


def nudge_right_output_if_overlaps():
    """
    If the right output block would overlap the left output block, start the right
    block in the first column after the left block (so e.g. left A–F → right can start at G).
    """
    global RIGHT_OUTPUT_START_COL
    left_w = block_width_cols(LEFT_BLOCK_START_COL, LEFT_BLOCK_END_COL)
    left_out_start = col_to_num(LEFT_OUTPUT_START_COL)
    left_out_end = left_out_start + left_w - 1
    right_out_start = col_to_num(RIGHT_OUTPUT_START_COL)
    if right_out_start <= left_out_end:
        RIGHT_OUTPUT_START_COL = get_column_letter(left_out_end + 1)


def sanitize_excel_sheet_title(title, default="Aligned"):
    """
    Excel worksheet names max 31 chars and cannot contain: \\ / ? * [ ]
    """
    raw = (title or "").strip()
    if not raw:
        raw = default
    illegal = '\\/*?:[]'
    cleaned = "".join(c for c in raw if c not in illegal).strip()
    if not cleaned:
        cleaned = default
    return cleaned[:31]


def highlight_block(ws, row, start_col, end_col):
    """
    Highlight a row block yellow.
    """
    start = col_to_num(start_col)
    end = col_to_num(end_col)

    for col in range(start, end + 1):
        ws.cell(row=row, column=col).fill = UNMATCHED_FILL


def highlight_difference_cell(ws, row, diff_col):
    """
    Highlight only the configured difference cell.
    """
    d = col_to_num(diff_col)
    ws.cell(row=row, column=d).fill = DIFF_ROW_FILL


def read_column(ws, col_letter):
    """
    Read one matching column into a list.

    Each item stores:
    - original Excel row
    - numeric value used for matching
    """
    rows = []

    for row in range(START_ROW, ws.max_row + 1):
        value = to_float(ws[f"{col_letter}{row}"].value)

        if value is not None:
            rows.append({
                "row": row,
                "value": value
            })

    return rows


def is_match(left_value, right_value):
    """
    Values match if their absolute difference is within threshold.
    """
    return abs(left_value - right_value) <= THRESHOLD


def pair_score(left_item, right_item):
    """
    Score for aligning one left value with one right value.
    """
    if is_match(left_item["value"], right_item["value"]):
        return MATCH_SCORE

    return MISMATCH_SCORE


# =========================
# Dynamic Programming
# =========================

def align_with_dp(left, right):
    """
    Full DP sequence alignment.

    left  = values from column D
    right = values from column F

    DP choices:
    1. DIAG = align left[i] with right[j]
    2. UP   = left[i] has no matching right value
    3. LEFT = right[j] has no matching left value
    """

    n = len(left)
    m = len(right)

    dp = [[0] * (m + 1) for _ in range(n + 1)]
    trace = [[None] * (m + 1) for _ in range(n + 1)]

    # First column: left values aligned with blanks
    for i in range(1, n + 1):
        dp[i][0] = dp[i - 1][0] + GAP_SCORE
        trace[i][0] = "UP"

    # First row: right values aligned with blanks
    for j in range(1, m + 1):
        dp[0][j] = dp[0][j - 1] + GAP_SCORE
        trace[0][j] = "LEFT"

    # Fill DP table
    for i in range(1, n + 1):
        for j in range(1, m + 1):
            left_item = left[i - 1]
            right_item = right[j - 1]

            diag = dp[i - 1][j - 1] + pair_score(left_item, right_item)
            up = dp[i - 1][j] + GAP_SCORE
            left_gap = dp[i][j - 1] + GAP_SCORE

            best = max(diag, up, left_gap)
            dp[i][j] = best

            # Prefer real matches.
            # Avoid mismatches unless they are truly best.
            if best == diag and is_match(left_item["value"], right_item["value"]):
                trace[i][j] = "DIAG"
            elif best == up:
                trace[i][j] = "UP"
            elif best == left_gap:
                trace[i][j] = "LEFT"
            else:
                trace[i][j] = "DIAG"

    # Backtrack from bottom-right
    alignment = []
    i = n
    j = m

    while i > 0 or j > 0:
        direction = trace[i][j]

        if direction == "DIAG":
            left_item = left[i - 1]
            right_item = right[j - 1]

            alignment.append({
                "left": left_item,
                "right": right_item,
                "matched": is_match(left_item["value"], right_item["value"])
            })

            i -= 1
            j -= 1

        elif direction == "UP":
            left_item = left[i - 1]

            alignment.append({
                "left": left_item,
                "right": None,
                "matched": False
            })

            i -= 1

        elif direction == "LEFT":
            right_item = right[j - 1]

            alignment.append({
                "left": None,
                "right": right_item,
                "matched": False
            })

            j -= 1

        else:
            break

    alignment.reverse()
    return alignment


# =========================
# Output Sheet
# =========================

def create_output_sheet(wb, src_ws, sheet_title):
    """
    Create a new output sheet for aligned results.
    Original sheet is not modified.
    """
    if sheet_title in wb.sheetnames:
        del wb[sheet_title]

    out_ws = wb.create_sheet(sheet_title)

    # Copy header rows from source onto the output at the same row numbers
    for hr in range(HEADER_FIRST_ROW, HEADER_LAST_ROW + 1):
        copy_block(
            src_ws,
            out_ws,
            hr,
            hr,
            LEFT_BLOCK_START_COL,
            LEFT_BLOCK_END_COL,
            LEFT_OUTPUT_START_COL,
        )
        copy_block(
            src_ws,
            out_ws,
            hr,
            hr,
            RIGHT_BLOCK_START_COL,
            RIGHT_BLOCK_END_COL,
            RIGHT_OUTPUT_START_COL,
        )

    # Copy column widths from source sheet
    for col_letter, dim in src_ws.column_dimensions.items():
        if dim.width is not None:
            out_ws.column_dimensions[col_letter].width = dim.width

    return out_ws


def write_alignment(src_ws, out_ws, alignment):
    """
    Write aligned full-row blocks into the output sheet.

    Left block:
        A:D

    Right block:
        F:H

    Left and right widths come from config; leave empty columns between them on
    the output sheet if their start columns leave a gap.
    """
    output_row = START_ROW

    for step in alignment:
        left_item = step["left"]
        right_item = step["right"]
        matched = step["matched"]

        # Copy left row block A:D
        if left_item is not None:
            copy_block(
                src_ws,
                out_ws,
                left_item["row"],
                output_row,
                LEFT_BLOCK_START_COL,
                LEFT_BLOCK_END_COL,
                LEFT_OUTPUT_START_COL
            )

        # Copy right row block F:H
        if right_item is not None:
            copy_block(
                src_ws,
                out_ws,
                right_item["row"],
                output_row,
                RIGHT_BLOCK_START_COL,
                RIGHT_BLOCK_END_COL,
                RIGHT_OUTPUT_START_COL
            )

        # Highlight unmatched row blocks
        if not matched:
            if left_item is not None:
                highlight_block(
                    out_ws,
                    output_row,
                    LEFT_OUTPUT_START_COL,
                    LEFT_BLOCK_END_COL
                )

            if right_item is not None:
                highlight_block(
                    out_ws,
                    output_row,
                    RIGHT_OUTPUT_START_COL,
                    RIGHT_BLOCK_END_COL
                )

        output_row += 1


def write_difference_column(out_ws, alignment, diff_col):
    """
    After rows are pasted, fill one column with abs(left − right) for the match values.

    Rows with only left or only right aligned stay blank in the diff column.
    """
    d = col_to_num(diff_col)
    out_ws.cell(row=HEADER_LAST_ROW, column=d).value = "Abs diff"

    output_row = START_ROW
    for step in alignment:
        left_item = step["left"]
        right_item = step["right"]
        if left_item is not None and right_item is not None:
            out_ws.cell(row=output_row, column=d).value = abs(
                left_item["value"] - right_item["value"]
            )
            highlight_difference_cell(out_ws, output_row, diff_col)
        else:
            out_ws.cell(row=output_row, column=d).value = None

        output_row += 1


# =========================
# Web-App Wrapper / Main
# =========================

def process_excel(
    input_file,
    output_file,
    input_sheet_name="sheet1",
    output_sheet_name="sheet 1 aligned result",
    start_row=2,
    header_first_row=1,
    header_last_row=1,
    left_input_col="D",
    left_block_start_col="A",
    left_block_end_col="D",
    left_output_start_col="A",
    right_input_col="F",
    right_block_start_col="F",
    right_block_end_col="H",
    right_output_start_col="F",
    threshold=0.5,
    diff_output_col="E",
):
    """
    Main processing function.

    This is what the web app will call.
    It takes user-entered settings instead of relying only on hardcoded values.
    """

    global INPUT_FILE, OUTPUT_FILE
    global INPUT_SHEET_NAME, OUTPUT_SHEET_NAME
    global START_ROW, HEADER_FIRST_ROW, HEADER_LAST_ROW
    global LEFT_INPUT_COL, LEFT_BLOCK_START_COL, LEFT_BLOCK_END_COL, LEFT_OUTPUT_START_COL
    global RIGHT_INPUT_COL, RIGHT_BLOCK_START_COL, RIGHT_BLOCK_END_COL, RIGHT_OUTPUT_START_COL
    global THRESHOLD

    INPUT_FILE = input_file
    OUTPUT_FILE = output_file

    INPUT_SHEET_NAME = input_sheet_name
    resolved_output_sheet = sanitize_excel_sheet_title(output_sheet_name)
    OUTPUT_SHEET_NAME = resolved_output_sheet

    START_ROW = int(start_row)
    hf = int(header_first_row)
    hl = int(header_last_row)
    if hl < hf:
        raise ValueError(
            "Header last row must be greater than or equal to header first row."
        )
    HEADER_FIRST_ROW = hf
    HEADER_LAST_ROW = hl
    if START_ROW <= HEADER_LAST_ROW:
        raise ValueError(
            "Data row must be below all header rows (after header last row)."
        )

    LEFT_INPUT_COL = left_input_col.upper()
    LEFT_BLOCK_START_COL = left_block_start_col.upper()
    LEFT_BLOCK_END_COL = left_block_end_col.upper()
    LEFT_OUTPUT_START_COL = left_output_start_col.upper()

    RIGHT_INPUT_COL = right_input_col.upper()
    RIGHT_BLOCK_START_COL = right_block_start_col.upper()
    RIGHT_BLOCK_END_COL = right_block_end_col.upper()
    RIGHT_OUTPUT_START_COL = right_output_start_col.upper()

    THRESHOLD = float(threshold)

    assert_match_column_in_block(
        LEFT_INPUT_COL, LEFT_BLOCK_START_COL, LEFT_BLOCK_END_COL, "First column group"
    )
    assert_match_column_in_block(
        RIGHT_INPUT_COL, RIGHT_BLOCK_START_COL, RIGHT_BLOCK_END_COL, "Second column group"
    )
    nudge_right_output_if_overlaps()

    diff_clean = (diff_output_col or "").strip().upper()
    if diff_clean:
        try:
            col_to_num(diff_clean)
        except Exception:
            raise ValueError(
                "Difference column must be valid letters (examples: E, AA)."
            ) from None
        assert_diff_column_clear_of_blocks(diff_clean)

    wb = load_workbook(INPUT_FILE)

    if INPUT_SHEET_NAME in wb.sheetnames:
        src_ws = wb[INPUT_SHEET_NAME]
    else:
        src_ws = wb.active
        print(f"Warning: sheet '{INPUT_SHEET_NAME}' was not found.")
        print(f"Using active sheet instead: '{src_ws.title}'")

    left = read_column(src_ws, LEFT_INPUT_COL)
    right = read_column(src_ws, RIGHT_INPUT_COL)

    alignment = align_with_dp(left, right)

    out_ws = create_output_sheet(wb, src_ws, resolved_output_sheet)
    write_alignment(src_ws, out_ws, alignment)
    if diff_clean:
        write_difference_column(out_ws, alignment, diff_clean)

    wb.save(OUTPUT_FILE)

    result = {
        "output_file": str(OUTPUT_FILE),
        "input_sheet": src_ws.title,
        "output_sheet": OUTPUT_SHEET_NAME,
        "left_values": len(left),
        "right_values": len(right),
        "alignment_steps": len(alignment),
        "matches_written": sum(1 for step in alignment if step["matched"]),
    }

    print("Done:", result["output_file"])
    print("Input sheet:", result["input_sheet"])
    print("Output sheet:", result["output_sheet"])
    print("Left values:", result["left_values"])
    print("Right values:", result["right_values"])
    print("Alignment steps:", result["alignment_steps"])
    print("Matches written:", result["matches_written"])

    return result


def main():
    """
    Allows you to still run this file directly from terminal:

        python3 version1.py
    """
    process_excel(
        input_file=INPUT_FILE,
        output_file=OUTPUT_FILE,
        input_sheet_name=INPUT_SHEET_NAME,
        output_sheet_name=OUTPUT_SHEET_NAME,
        start_row=START_ROW,
        header_first_row=HEADER_FIRST_ROW,
        header_last_row=HEADER_LAST_ROW,
        left_input_col=LEFT_INPUT_COL,
        left_block_start_col=LEFT_BLOCK_START_COL,
        left_block_end_col=LEFT_BLOCK_END_COL,
        left_output_start_col=LEFT_OUTPUT_START_COL,
        right_input_col=RIGHT_INPUT_COL,
        right_block_start_col=RIGHT_BLOCK_START_COL,
        right_block_end_col=RIGHT_BLOCK_END_COL,
        right_output_start_col=RIGHT_OUTPUT_START_COL,
        threshold=THRESHOLD,
        diff_output_col=DIFF_OUTPUT_COL,
    )


if __name__ == "__main__":
    main()