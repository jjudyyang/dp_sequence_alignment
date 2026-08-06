#!/usr/bin/env python3
"""Extract a compact workbook preview for the Shiftline frontend."""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def color_to_hex(color: Any) -> str | None:
    """Return a browser-friendly hex color when openpyxl exposes a direct RGB."""
    if not color:
        return None
    if color.type == "rgb" and color.rgb:
        raw = color.rgb.upper()
        if raw in {"00000000", "FFFFFFFF"}:
            return None if raw == "00000000" else "#FFFFFF"
        if len(raw) == 8:
            raw = raw[2:]
        return f"#{raw}"
    if color.type == "indexed":
        return f"indexed:{color.indexed}"
    if color.type == "theme":
        return f"theme:{color.theme}"
    return None


def safe_value(value: Any) -> Any:
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    return value


def style_snapshot(cell: Any) -> dict[str, Any]:
    fill = cell.fill
    font = cell.font
    alignment = cell.alignment
    border = cell.border

    border_summary = {}
    for edge in ("left", "right", "top", "bottom"):
        side = getattr(border, edge)
        if side and side.style:
            border_summary[edge] = {
                "style": side.style,
                "color": color_to_hex(side.color),
            }

    fill_color = None
    if fill and fill.fill_type:
        fill_color = color_to_hex(fill.fgColor) or color_to_hex(fill.start_color)

    return {
        "fill": {
            "type": fill.fill_type,
            "color": fill_color,
        },
        "font": {
            "name": font.name,
            "size": font.sz,
            "bold": bool(font.bold),
            "italic": bool(font.italic),
            "underline": font.underline,
            "color": color_to_hex(font.color),
        },
        "alignment": {
            "horizontal": alignment.horizontal,
            "vertical": alignment.vertical,
            "wrapText": bool(alignment.wrap_text),
        },
        "numberFormat": cell.number_format,
        "border": border_summary,
    }


def style_key(style: dict[str, Any]) -> str:
    packed = json.dumps(style, sort_keys=True, default=str)
    return hashlib.sha1(packed.encode("utf-8")).hexdigest()[:10]


def cell_payload(cell: Any, key: str) -> dict[str, Any]:
    return {
        "address": cell.coordinate,
        "row": cell.row,
        "col": cell.column,
        "value": safe_value(cell.value),
        "dataType": cell.data_type,
        "numberFormat": cell.number_format,
        "styleKey": key,
    }


def serialize_workbook(
    workbook_path: Path,
    max_rows: int,
    max_cols: int,
    source_url: str | None,
) -> dict[str, Any]:
    wb = load_workbook(workbook_path, data_only=False)

    styles: dict[str, dict[str, Any]] = {}
    sheets = []

    for ws in wb.worksheets:
        preview_rows = []
        col_count = min(ws.max_column, max_cols)
        row_count = min(ws.max_row, max_rows)

        for row_idx in range(1, row_count + 1):
            cells = []
            for col_idx in range(1, col_count + 1):
                cell = ws.cell(row_idx, col_idx)
                snapshot = style_snapshot(cell)
                key = style_key(snapshot)
                styles.setdefault(key, snapshot)
                cells.append(cell_payload(cell, key))
            preview_rows.append({"rowNumber": row_idx, "cells": cells})

        column_widths = []
        for col_idx in range(1, col_count + 1):
            letter = get_column_letter(col_idx)
            column_widths.append(
                {
                    "col": letter,
                    "width": ws.column_dimensions[letter].width,
                }
            )

        sheets.append(
            {
                "name": ws.title,
                "maxRows": ws.max_row,
                "maxCols": ws.max_column,
                "previewRows": preview_rows,
                "previewCols": [get_column_letter(i) for i in range(1, col_count + 1)],
                "columnWidths": column_widths,
                "freezePanes": str(ws.freeze_panes) if ws.freeze_panes else None,
            }
        )

    return {
        "generatedAt": dt.datetime.now(dt.timezone.utc).isoformat(),
        "source": {
            "fileName": workbook_path.name,
            "path": str(workbook_path),
            "driveUrl": source_url,
        },
        "workbook": {
            "sheetCount": len(wb.worksheets),
            "sheets": [
                {"name": ws.title, "maxRows": ws.max_row, "maxCols": ws.max_column}
                for ws in wb.worksheets
            ],
        },
        "sheets": sheets,
        "styles": styles,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("frontend/data/sample-workbook.js"),
    )
    parser.add_argument("--max-rows", type=int, default=80)
    parser.add_argument("--max-cols", type=int, default=12)
    parser.add_argument("--source-url", default=None)
    args = parser.parse_args()

    payload = serialize_workbook(
        args.workbook,
        max_rows=args.max_rows,
        max_cols=args.max_cols,
        source_url=args.source_url,
    )

    args.output.parent.mkdir(parents=True, exist_ok=True)
    body = json.dumps(payload, indent=2, ensure_ascii=True)
    args.output.write_text(f"window.WELD_SHIFT_SAMPLE = {body};\n", encoding="utf-8")
    print(f"Wrote {args.output}")
    print(
        f"Sheets: {len(payload['sheets'])}; preview styles: {len(payload['styles'])}"
    )


if __name__ == "__main__":
    main()
