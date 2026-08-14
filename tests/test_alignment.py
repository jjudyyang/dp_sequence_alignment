import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from alignment import AlignmentConfig, align_with_dp, process_excel


def sequence(values):
    return [{"row": index + 2, "value": value} for index, value in enumerate(values)]


class AlignmentTests(unittest.TestCase):
    def test_align_with_dp_preserves_order_and_inserts_gap(self):
        left = sequence([10.0, 20.0, 30.0])
        right = sequence([10.1, 30.2])

        alignment = align_with_dp(left, right, threshold=0.25)

        self.assertEqual(len(alignment), 3)
        self.assertTrue(alignment[0]["matched"])
        self.assertEqual(alignment[0]["left"]["value"], 10.0)
        self.assertEqual(alignment[0]["right"]["value"], 10.1)
        self.assertFalse(alignment[1]["matched"])
        self.assertEqual(alignment[1]["left"]["value"], 20.0)
        self.assertIsNone(alignment[1]["right"])
        self.assertTrue(alignment[2]["matched"])
        self.assertEqual(alignment[2]["left"]["value"], 30.0)
        self.assertEqual(alignment[2]["right"]["value"], 30.2)

    def test_process_excel_writes_shifted_workbook_from_two_sheets(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_path = tmp_path / "input.xlsx"
            output_path = tmp_path / "output.xlsx"

            wb = Workbook()
            left_ws = wb.active
            left_ws.title = "LeftData"
            left_ws["A1"] = "Part"
            left_ws["B1"] = "Distance"
            left_ws.append(["L1", 10.0])
            left_ws.append(["L2", 20.0])
            left_ws.append(["L3", 30.0])

            right_ws = wb.create_sheet("RightData")
            right_ws["D1"] = "Distance"
            right_ws["E1"] = "Label"
            right_ws["D2"] = 10.1
            right_ws["E2"] = "R1"
            right_ws["D3"] = 30.2
            right_ws["E3"] = "R3"
            wb.save(input_path)

            result = process_excel(
                input_file=input_path,
                output_file=output_path,
                input_sheet_name="LeftData",
                left_sheet_name="LeftData",
                right_sheet_name="RightData",
                output_sheet_name="Aligned",
                start_row=2,
                header_first_row=1,
                header_last_row=1,
                left_input_col="B",
                left_block_start_col="A",
                left_block_end_col="B",
                left_output_start_col="A",
                right_input_col="D",
                right_block_start_col="D",
                right_block_end_col="E",
                right_output_start_col="D",
                threshold=0.25,
                diff_output_col="C",
            )

            self.assertEqual(result["left_sheet"], "LeftData")
            self.assertEqual(result["right_sheet"], "RightData")
            self.assertEqual(result["alignment_steps"], 3)
            self.assertEqual(result["matches_written"], 2)

            out_wb = load_workbook(output_path)
            out_ws = out_wb["Aligned"]

            self.assertEqual(out_ws["A1"].value, "Part")
            self.assertEqual(out_ws["B1"].value, "Distance")
            self.assertEqual(out_ws["C1"].value, "Abs diff")
            self.assertEqual(out_ws["D1"].value, "Distance")
            self.assertEqual(out_ws["E1"].value, "Label")

            self.assertEqual(out_ws["A2"].value, "L1")
            self.assertEqual(out_ws["B2"].value, 10.0)
            self.assertAlmostEqual(out_ws["C2"].value, 0.1)
            self.assertEqual(out_ws["D2"].value, 10.1)
            self.assertEqual(out_ws["E2"].value, "R1")

            self.assertEqual(out_ws["A3"].value, "L2")
            self.assertEqual(out_ws["B3"].value, 20.0)
            self.assertIsNone(out_ws["C3"].value)
            self.assertIsNone(out_ws["D3"].value)
            self.assertIsNone(out_ws["E3"].value)

            self.assertEqual(out_ws["A4"].value, "L3")
            self.assertEqual(out_ws["B4"].value, 30.0)
            self.assertAlmostEqual(out_ws["C4"].value, 0.2)
            self.assertEqual(out_ws["D4"].value, 30.2)
            self.assertEqual(out_ws["E4"].value, "R3")

    def test_diff_column_cannot_overlap_output_blocks(self):
        with self.assertRaisesRegex(ValueError, "Difference column B overlaps"):
            AlignmentConfig(
                left_block_start_col="A",
                left_block_end_col="B",
                left_output_start_col="A",
                left_input_col="B",
                right_block_start_col="D",
                right_block_end_col="E",
                right_output_start_col="D",
                right_input_col="D",
                diff_output_col="B",
            )

    def test_process_excel_rejects_empty_match_column(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_path = tmp_path / "input.xlsx"
            output_path = tmp_path / "output.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            ws["D1"] = "Left distance"
            ws["F1"] = "Right distance"
            ws["D2"] = 10.0
            ws["D3"] = 20.0
            wb.save(input_path)

            with self.assertRaisesRegex(
                ValueError,
                "No numeric values found in right match column F on sheet 'Data'",
            ):
                process_excel(
                    input_file=input_path,
                    output_file=output_path,
                    input_sheet_name="Data",
                    left_sheet_name="Data",
                    right_sheet_name="Data",
                    output_sheet_name="Aligned",
                    start_row=2,
                    header_first_row=1,
                    header_last_row=1,
                    left_input_col="D",
                    left_block_start_col="A",
                    left_block_end_col="D",
                    left_output_start_col="A",
                    right_input_col="F",
                    right_block_start_col="F",
                    right_block_end_col="H",
                    right_output_start_col="F",
                    threshold=0.25,
                    diff_output_col="E",
                )

    def test_process_excel_infers_current_previous_single_sheet_layout(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_path = tmp_path / "input.xlsx"
            output_path = tmp_path / "output.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "sheet1"
            ws.append([None, "Current", None, None, None, None, None, None, "Previous"])
            ws.append(
                [
                    "Long Seam",
                    "Pipe Tpe",
                    "WT",
                    "Joint #",
                    "ID",
                    "Distance [ft]",
                    "Target Joint Length [ft]",
                    "Joint Length Dif [ft]",
                    "Joint Length [ft]",
                    "Distance [ft]",
                    "Joint Number",
                    "WT",
                    "Pipe Tipe",
                    "Long seam",
                    "HCA",
                ]
            )
            ws.append([None, "Valve", 0.219, None, 1, -1.575, 3.118, None, 3.128])
            ws.append([None, "Tee", None, None, 453, 1.543, 1.154, None, 1.2])
            wb.save(input_path)

            result = process_excel(
                input_file=input_path,
                output_file=output_path,
                input_sheet_name="sheet1",
                output_sheet_name="Aligned",
                threshold=0.25,
            )

            self.assertEqual(result["left_values"], 2)
            self.assertEqual(result["right_values"], 2)
            self.assertEqual(result["matches_written"], 2)

            out_wb = load_workbook(output_path)
            out_ws = out_wb["Aligned"]
            self.assertEqual(out_ws.max_column, 15)
            self.assertEqual(out_ws["G2"].value, "Target Joint Length [ft]")
            self.assertEqual(out_ws["H2"].value, "Abs diff")
            self.assertEqual(out_ws["I2"].value, "Joint Length [ft]")
            self.assertAlmostEqual(out_ws["H3"].value, 0.01)
            self.assertAlmostEqual(out_ws["H4"].value, 0.046)


if __name__ == "__main__":
    unittest.main()
